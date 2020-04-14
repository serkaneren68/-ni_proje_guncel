from openpyxl import load_workbook
from openpyxl import Workbook
def trans(x):
    a = x.replace(" ","_");
    b = a.lower()
    return b
def convert(x):
    if("(" in x):
        index = x.find("(")
        if(x[index-1] == " "):

            a = x[0:index-1]
            return a
        else:
            a = x[0:index]
            return a
    else:
        return x[0:-1] ;



workbook = load_workbook("toplu2.xlsx")

sheet = workbook.active

k = 1
j = 1
bolumler = list()

for i in range(1,11403):

    if not(convert(sheet["B{}".format(i)].value) in bolumler) :
        bolumler.append(convert(sheet["B{}".format(i)].value))
        wb = Workbook()
        ws = wb.active
        ws.append(["Üniversite","Bölüm","Üni. Türü/Ücret","Kontenjan","Taban Başarı Sırası(0.12)\n2019\n2018\n2017\n2016","Taban Puanı(0.12)\n2019\n2018\n2017\n2016"])
        wb.save("lisansv4/{}.xlsx".format(trans(bolumler[k-1])))
        l = k
        k = k+1


    if (convert(sheet["B{}".format(i)].value) in bolumler):
        wb = load_workbook("lisansv4/{}.xlsx".format(trans(bolumler[l-1])))
        ws = wb.active

        ws.append([sheet["A{}".format(i)].value,sheet["B{}".format(i)].value,sheet["C{}".format(i)].value,sheet["D{}".format(i)].value,sheet["E{}".format(i)].value,sheet["F{}".format(i)].value])
        wb.save("lisansv4/{}.xlsx".format(trans(bolumler[l-1])))


for i in bolumler:
    print(i)
